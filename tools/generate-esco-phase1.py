#!/usr/bin/env python3
# ============================================================================
# generate-esco-phase1.py — генератор SQL-міграції Phase 1 для тенанта 'esco'.
# ============================================================================
#
# ЩО РОБИТЬ
# ---------
# Читає три xlsx з /PerenosTablu/:
#   Config_crm_v2.xlsx   → Налаштування, Власник, Персонал
#   Kliyentu_crm_v3.xlsx → Клієнти (clients_directory), Сповіщення, Контент апки
#   Finance_crm_v2.xlsx  → Шаблон розподілу
#
# Генерує один SQL-файл sql/2026-04-esco-phase1-foundation.sql з:
#   1. ALTER TABLE owner_account ADD COLUMN IF NOT EXISTS legacy_id, access_scope, notes.
#   2. CREATE UNIQUE INDEX IF NOT EXISTS для (tenant_id, <legacy_id>) на 6 таблицях.
#   3. INSERT ... ON CONFLICT DO NOTHING для ~57 рядків.
#
# tenant_id скрізь фіксовано = 'esco'.
# Мова 'UA' → 'uk' (нормалізація).
# Ролі: 'Водій'→'driver', 'Менеджер'→'manager', 'Власник'→'owner'.
# Булеві: 'Так'→true, 'Ні'→false, None→NULL.
# Паролі: NULL скрізь окрім owner_account, де NOT NULL — ставимо
#   'CHANGE_ME_ESCO_<login>' як плейсхолдер.
#
# ВИКОРИСТАННЯ
# -----------
#   python3 tools/generate-esco-phase1.py
# Вихід: sql/2026-04-esco-phase1-foundation.sql
# ============================================================================

import openpyxl
import os
import re
import sys
from datetime import datetime, date

TENANT = 'esco'
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, 'sql', '2026-04-esco-phase1-foundation.sql')

# ── Helpers: SQL-escape і типізація ─────────────────────────────────────────

def q(v):
    """Quote value as SQL literal. None → NULL."""
    if v is None:
        return 'NULL'
    if isinstance(v, bool):
        return 'true' if v else 'false'
    if isinstance(v, (int, float)):
        # не округляємо; if float is integral, віддаємо як int
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v)
    if isinstance(v, datetime):
        return "'" + v.strftime('%Y-%m-%d %H:%M:%S') + "'::timestamptz"
    if isinstance(v, date):
        return "'" + v.strftime('%Y-%m-%d') + "'::date"
    s = str(v).replace("'", "''")
    return "'" + s + "'"


def nn(v):
    """Normalize: порожній рядок / 'None' string → None."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if not s or s.lower() in ('none', 'null', '#n/a', '#value!'):
            return None
        return s
    return v


def to_bool(v):
    v = nn(v)
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in ('так', 'yes', 'true', 'активний', 'активно', '1', 'y'):
        return True
    if s in ('ні', 'no', 'false', '0', 'n'):
        return False
    return None


def to_num(v):
    v = nn(v)
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = re.sub(r'[^\d.\-]', '', str(v).replace(',', '.'))
    if not s or s in ('.', '-', '-.'):
        return None
    try:
        return float(s) if '.' in s else int(s)
    except ValueError:
        return None


def to_date(v):
    v = nn(v)
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v
    # try parse common formats
    s = str(v).strip()
    for fmt in ('%d.%m.%Y %H:%M', '%d.%m.%Y', '%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S.%fZ', '%Y-%m-%d'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def norm_phone(v):
    v = nn(v)
    if v is None:
        return None
    # xlsx іноді зберігає телефон як float (955222526.0) — прибираємо хвіст .0
    if isinstance(v, float) and v.is_integer():
        v = str(int(v))
    s = re.sub(r'[\s()\-]', '', str(v))
    # прибираємо «.0» що могло зʼїхати через str(float)
    if re.match(r'^\d+\.0$', s):
        s = s[:-2]
    if s.startswith('+'):
        return s
    if re.match(r'^380\d{9}$', s):
        return '+' + s
    if re.match(r'^0\d{9}$', s):
        return '+38' + s
    if re.match(r'^\d{10}$', s):
        return '+38' + s  # ймовірно UA без 0/380
    return s  # foreign/inknown — as-is


STATUS_MAP = {
    'активний': 'active',
    'активно': 'active',
    'неактивний': 'inactive',
    'неактивно': 'inactive',
    'заблокований': 'suspended',
    'заблоковано': 'suspended',
}

def norm_status(v, default='active'):
    v = nn(v)
    if v is None:
        return default
    key = str(v).strip().lower()
    return STATUS_MAP.get(key, key)


ROLE_MAP = {
    'водій': 'driver',
    'менеджер': 'manager',
    'власник': 'owner',
    'диспетчер': 'dispatcher',
}

def norm_role(v):
    v = nn(v)
    if v is None:
        return None
    key = str(v).strip().lower()
    return ROLE_MAP.get(key, key)


def norm_lang(v):
    v = nn(v)
    if v is None:
        return None
    s = str(v).strip().upper()
    return {'UA': 'uk', 'EN': 'en', 'DE': 'de', 'RU': 'ru', 'PL': 'pl'}.get(s, s.lower())


# ── Reading xlsx sheets into list-of-dict ──────────────────────────────────

def read_sheet(path, sheet_name):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f'Sheet {sheet_name!r} not found in {path}')
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    out = []
    for r in rows[1:]:
        if all(v is None or v == '' for v in r):
            continue
        d = {}
        for i, header in enumerate(headers):
            d[header] = r[i] if i < len(r) else None
        out.append(d)
    wb.close()
    return out


# ── Build INSERT-set for each target table ─────────────────────────────────

def build_system_settings(rows):
    sql = []
    for r in rows:
        setting_id = nn(r.get('SETTING_ID'))
        if not setting_id:
            continue
        notes_parts = []
        if nn(r.get('Хто змінив')):
            notes_parts.append(f"Хто змінив: {nn(r.get('Хто змінив'))}")
        if nn(r.get('Примітка')):
            notes_parts.append(str(nn(r.get('Примітка'))))
        notes = '; '.join(notes_parts) if notes_parts else None

        sql.append(
            f"INSERT INTO public.system_settings "
            f"(tenant_id, setting_id, setting_section, setting_name, setting_value, setting_description, updated_date, notes) VALUES ("
            f"{q(TENANT)}, {q(setting_id)}, {q(nn(r.get('Розділ')))}, {q(nn(r.get('Параметр')))}, "
            f"{q(nn(r.get('Значення')))}, {q(nn(r.get('Опис')))}, {q(to_date(r.get('Дата зміни')))}, {q(notes)}"
            f") ON CONFLICT (tenant_id, setting_id) DO NOTHING;"
        )
    return sql


def build_app_content(rows):
    sql = []
    for r in rows:
        cid = nn(r.get('CONTENT_ID'))
        if not cid:
            continue
        notes_parts = []
        if nn(r.get('Хто змінив')):
            notes_parts.append(f"Хто змінив: {nn(r.get('Хто змінив'))}")
        if nn(r.get('Примітка')):
            notes_parts.append(str(nn(r.get('Примітка'))))
        notes = '; '.join(notes_parts) if notes_parts else None

        sql.append(
            f"INSERT INTO public.app_content "
            f"(tenant_id, content_id, content_type, content_name, language, content_text, is_active, updated_date, notes) VALUES ("
            f"{q(TENANT)}, {q(cid)}, {q(nn(r.get('Тип контенту')))}, {q(nn(r.get('Назва')))}, "
            f"{q(norm_lang(r.get('Мова')))}, {q(nn(r.get('Текст/Значення')))}, "
            f"{q(to_bool(r.get('Активний')))}, {q(to_date(r.get('Дата змін')))}, {q(notes)}"
            f") ON CONFLICT (tenant_id, content_id) DO NOTHING;"
        )
    return sql


def build_distribution_template(rows):
    sql = []
    for r in rows:
        tid = nn(r.get('TMPL_ID'))
        if not tid:
            continue
        notes_parts = []
        if nn(r.get('Отримувач')):
            notes_parts.append(f"Отримувач: {nn(r.get('Отримувач'))}")
        if nn(r.get('Примітка')):
            notes_parts.append(str(nn(r.get('Примітка'))))
        notes = '; '.join(notes_parts) if notes_parts else None

        sql.append(
            f"INSERT INTO public.distribution_template "
            f"(tenant_id, tmpl_id, category, description, default_percentage, fixed_amount, amount_currency, payment_form, is_active, notes) VALUES ("
            f"{q(TENANT)}, {q(tid)}, {q(nn(r.get('Категорія')))}, {q(nn(r.get('Опис')))}, "
            f"{q(to_num(r.get('% за замовч')))}, {q(to_num(r.get('Фіксована сума')))}, "
            f"{q(nn(r.get('Валюта')))}, {q(nn(r.get('Форма виплати')))}, "
            f"{q(to_bool(r.get('Активний')))}, {q(notes)}"
            f") ON CONFLICT (tenant_id, tmpl_id) DO NOTHING;"
        )
    return sql


def build_notifications(rows):
    sql = []
    for r in rows:
        nid = nn(r.get('NOTIF_ID'))
        if not nid:
            continue
        sql.append(
            f"INSERT INTO public.notifications "
            f"(tenant_id, notif_id, notification_type, event_trigger, channel, template_text, is_active, sent_at, delivery_status, sent_message_text) VALUES ("
            f"{q(TENANT)}, {q(nid)}, {q(nn(r.get('Тип')))}, {q(nn(r.get('Подія')))}, "
            f"{q(nn(r.get('Канал')))}, {q(nn(r.get('Шаблон тексту')))}, "
            f"{q(to_bool(r.get('Активний')))}, {q(to_date(r.get('Дата відправки')))}, "
            f"{q(nn(r.get('Статус відправки')))}, {q(nn(r.get('Текст відправлений')))}"
            f") ON CONFLICT (tenant_id, notif_id) DO NOTHING;"
        )
    return sql


def to_rating(v, max_val=5.0):
    """Рейтинг у межах (0, max_val]. Нуль означає «ще нема оцінок» → NULL.
    Якщо значення поза межами — повертає (None, raw) для збереження raw у notes."""
    n = to_num(v)
    if n is None:
        return (None, None)
    if n == 0:
        # 0 = «рейтингу ще нема», у БД CHECK constraint не пускає 0 — ставимо NULL,
        # нічого не губимо (ratings_count=0 все одно каже «немає оцінок»).
        return (None, None)
    if 0 < n <= max_val:
        return (n, None)
    return (None, n)  # поза діапазоном → у БД NULL, raw → у notes


def build_clients_directory(rows):
    sql = []
    for r in rows:
        cid = nn(r.get('CLI_ID'))
        if not cid:
            continue

        # Рейтинги з CHECK-constraint 0..5. Якщо raw поза межами — зберігаємо в notes.
        driver_rating,   raw_drv = to_rating(r.get('Рейт. водія'))
        manager_rating,  raw_mgr = to_rating(r.get('Рейт. менеджера'))
        internal_rating, raw_int = to_rating(r.get('Внутрішній рейтинг'))
        bot_rating,      raw_bot = to_rating(r.get('Рейт. через бот'))

        raw_notes = []
        if raw_drv is not None:   raw_notes.append(f"driver_rating(raw)={raw_drv}")
        if raw_mgr is not None:   raw_notes.append(f"manager_rating(raw)={raw_mgr}")
        if raw_int is not None:   raw_notes.append(f"internal_rating(raw)={raw_int}")
        if raw_bot is not None:   raw_notes.append(f"bot_rating(raw)={raw_bot}")
        notes_parts = []
        if raw_notes:
            notes_parts.append('OUT_OF_RANGE: ' + ', '.join(raw_notes))
        if nn(r.get('Примітка')):
            notes_parts.append(str(nn(r.get('Примітка'))))
        notes = '; '.join(notes_parts) if notes_parts else None

        sql.append(
            f"INSERT INTO public.clients_directory ("
            f"tenant_id, cli_id, smart_id, registered_at, last_activity, phone, additional_phone, "
            f"full_name, email, direction, client_type, is_vip, in_stoplist, stoplist_reason, "
            f"total_routes, total_packages, total_bookings, last_payment_date, "
            f"debt_uah, debt_chf, debt_eur, debt_pln, debt_czk, "
            f"driver_rating, driver_ratings_count, driver_total_score, "
            f"manager_rating, manager_ratings_count, manager_total_score, internal_rating, "
            f"bot_rating, bot_ratings_count, bot_total_score, "
            f"emoji_positive, emoji_neutral, emoji_negative, "
            f"last_three_comments, last_review, last_review_date, app_status, notes) VALUES ("
            f"{q(TENANT)}, {q(cid)}, {q(nn(r.get('Ід_смарт/CRM')))}, "
            f"{q(to_date(r.get('Дата реєстрації')))}, {q(to_date(r.get('Остання активність')))}, "
            f"{q(norm_phone(r.get('Телефон')))}, {q(norm_phone(r.get('Додатковий телефон')))}, "
            f"{q(nn(r.get('Піб')))}, {q(nn(r.get('EMAIL')))}, "
            f"{q(nn(r.get('Напрям')))}, {q(nn(r.get('Тип клієнта')))}, "
            f"{q(to_bool(r.get('VIP')))}, {q(to_bool(r.get('Стоп-лист')))}, {q(nn(r.get('Причина стоп-листа')))}, "
            f"{q(to_num(r.get('К-сть рейсів')))}, {q(to_num(r.get('К-сть посилок')))}, {q(to_num(r.get('К-сть бронювань')))}, "
            f"{q(to_date(r.get('Остання оплата')))}, "
            f"{q(to_num(r.get('Борг UAH')))}, {q(to_num(r.get('Борг CHF')))}, {q(to_num(r.get('Борг EUR')))}, "
            f"{q(to_num(r.get('Борг PLN')))}, {q(to_num(r.get('Борг CZK')))}, "
            f"{q(driver_rating)}, {q(to_num(r.get('Оцінок від водія')))}, {q(to_num(r.get('Сума балів водія')))}, "
            f"{q(manager_rating)}, {q(to_num(r.get('Оцінок від менеджера')))}, {q(to_num(r.get('Сума балів менеджера')))}, "
            f"{q(internal_rating)}, "
            f"{q(bot_rating)}, {q(to_num(r.get('Оцінок через бот')))}, {q(to_num(r.get('Сума балів бот')))}, "
            f"{q(to_num(r.get('Супер 😊')))}, {q(to_num(r.get('Добре 😐')))}, {q(to_num(r.get('Погано 😞')))}, "
            f"{q(nn(r.get('Останні 3 коментарі')))}, {q(nn(r.get('Останній відгук')))}, "
            f"{q(to_date(r.get('Дата останнього відгуку')))}, "
            f"{q(norm_status(r.get('Статус апки')))}, {q(notes)}"
            f") ON CONFLICT (tenant_id, cli_id) DO NOTHING;"
        )
    return sql


def build_staff(rows):
    sql = []
    for r in rows:
        sid = nn(r.get('STAFF_ID'))
        if not sid:
            continue
        # AUTO_ID/Номер авто → у notes, бо vehicle_id (uuid) буде заповнений у Phase 2
        notes_parts = []
        if nn(r.get('AUTO_ID')) or nn(r.get('Номер авто')):
            notes_parts.append(f"Авто: {nn(r.get('AUTO_ID')) or '-'} / {nn(r.get('Номер авто')) or '-'}")
        if nn(r.get('Примітка')):
            notes_parts.append(str(nn(r.get('Примітка'))))
        notes = '; '.join(notes_parts) if notes_parts else None

        role = norm_role(r.get('Роль')) or 'staff'  # NOT NULL у таблиці
        sql.append(
            f"INSERT INTO public.staff ("
            f"tenant_id, staff_id, full_name, phone, email, role, login, "
            f"city_based, salary, salary_currency, employment_status, employment_date, last_activity, notes) VALUES ("
            f"{q(TENANT)}, {q(sid)}, {q(nn(r.get('Піб')))}, "
            f"{q(norm_phone(r.get('Телефон')))}, {q(nn(r.get('EMAIL')))}, {q(role)}, "
            f"{q(nn(r.get('Логін')))}, {q(nn(r.get('Місто базування')))}, "
            f"{q(to_num(r.get('Ставка')))}, {q(nn(r.get('Валюта ставки')))}, "
            f"{q(norm_status(r.get('Статус')))}, {q(to_date(r.get('Дата прийому')))}, "
            f"{q(to_date(r.get('Остання активність')))}, {q(notes)}"
            f") ON CONFLICT (tenant_id, staff_id) DO NOTHING;"
        )
    return sql


def build_owner_account(rows):
    """owner_account має UNIQUE(tenant_id) — один тенант = один власник.
    Вставляємо ТІЛЬКИ ПЕРШОГО власника (primary). Інші йдуть у staff як
    role='owner' через build_staff (їх додає extra_owners параметр)."""
    sql = []
    for r in rows[:1]:  # тільки primary
        login = nn(r.get('Логін'))
        if not login:
            continue
        pwd_placeholder = f'CHANGE_ME_ESCO_{login}'
        access_scope = nn(r.get('Доступ до таблиць'))
        notes = nn(r.get('Примітка'))

        sql.append(
            f"INSERT INTO public.owner_account ("
            f"tenant_id, legacy_id, full_name, phone, email, login, password_hash, "
            f"api_token, two_fa_enabled, account_status, created_at, last_login, last_password_change, "
            f"access_scope, notes) VALUES ("
            f"{q(TENANT)}, {q(nn(r.get('USER_ID')))}, {q(nn(r.get('Піб')))}, "
            f"{q(norm_phone(r.get('Телефон')))}, {q(nn(r.get('EMAIL')))}, "
            f"{q(login)}, {q(pwd_placeholder)}, "
            f"{q(nn(r.get('Токен API')))}, {q(to_bool(r.get('2FA активно')))}, "
            f"{q(norm_status(r.get('Статус')))}, "
            f"{q(to_date(r.get('Дата створення')))}, {q(to_date(r.get('Остання активність')))}, "
            f"{q(to_date(r.get('Дата зміни пароля')))}, "
            f"{q(access_scope)}, {q(notes)}"
            f") ON CONFLICT (tenant_id) DO NOTHING;"
        )
    return sql


def build_extra_owners_as_staff(owner_rows):
    """Додаткові власники (не primary) йдуть у staff з role='owner',
    щоб не втратити їхні дані. primary = перший у списку."""
    sql = []
    for r in owner_rows[1:]:
        uid = nn(r.get('USER_ID'))
        if not uid:
            continue
        notes_parts = ['Другий власник (primary — у owner_account)']
        if nn(r.get('Доступ до таблиць')):
            notes_parts.append(f"Доступ: {nn(r.get('Доступ до таблиць'))}")
        if nn(r.get('Примітка')):
            notes_parts.append(str(nn(r.get('Примітка'))))
        notes = '; '.join(notes_parts)

        sql.append(
            f"INSERT INTO public.staff ("
            f"tenant_id, staff_id, full_name, phone, email, role, login, "
            f"employment_status, last_activity, notes) VALUES ("
            f"{q(TENANT)}, {q(uid)}, {q(nn(r.get('Піб')))}, "
            f"{q(norm_phone(r.get('Телефон')))}, {q(nn(r.get('EMAIL')))}, "
            f"{q('owner')}, {q(nn(r.get('Логін')))}, "
            f"{q(norm_status(r.get('Статус')))}, "
            f"{q(to_date(r.get('Остання активність')))}, {q(notes)}"
            f") ON CONFLICT (tenant_id, staff_id) DO NOTHING;"
        )
    return sql


# ── Main ────────────────────────────────────────────────────────────────────

def main():
    config_xlsx = os.path.join(ROOT, 'PerenosTablu', 'Config_crm_v2.xlsx')
    clients_xlsx = os.path.join(ROOT, 'PerenosTablu', 'Kliyentu_crm_v3.xlsx')
    finance_xlsx = os.path.join(ROOT, 'PerenosTablu', 'Finance_crm_v2.xlsx')

    settings = read_sheet(config_xlsx, 'Налаштування')
    owners = read_sheet(config_xlsx, 'Власник')
    staff = read_sheet(config_xlsx, 'Персонал')

    clients = read_sheet(clients_xlsx, 'Клієнти')
    notifications = read_sheet(clients_xlsx, 'Сповіщення')
    app_content = read_sheet(clients_xlsx, 'Контент апки')

    templates = read_sheet(finance_xlsx, 'Шаблон розподілу')

    counts = {
        'system_settings': len(settings),
        'app_content': len(app_content),
        'distribution_template': len(templates),
        'notifications': len(notifications),
        'clients_directory': len(clients),
        'staff': len(staff),
        'owner_account': len(owners),
    }
    total = sum(counts.values())

    out = []
    out.append("-- =========================================================================")
    out.append("-- ESCO MIGRATION — PHASE 1: FOUNDATION")
    out.append("-- =========================================================================")
    out.append("--")
    out.append("-- АВТОГЕНЕРОВАНО tools/generate-esco-phase1.py — НЕ редагувати вручну.")
    out.append("-- Якщо треба виправити рядок — редагуй xlsx у PerenosTablu/ і перезапусти")
    out.append("-- скрипт, або ALTER/UPDATE окремим файлом поверх.")
    out.append("--")
    out.append("-- Переносимо до public.<table> з tenant_id='esco' ~{} рядків:".format(total))
    for k, v in counts.items():
        out.append("--   {:<25} {:>4} рядків".format(k, v))
    out.append("--")
    out.append("-- Ідемпотентно: ON CONFLICT (tenant_id, <legacy_id>) DO NOTHING. Повторний")
    out.append("-- запуск нічого не дублює.")
    out.append("-- =========================================================================")
    out.append("")
    out.append("BEGIN;")
    out.append("")

    # ── 1. ALTER owner_account: додаємо legacy_id, access_scope, notes ─────
    out.append("-- ── 1. Додаткові колонки на owner_account (Phase 1 рішення Q3=а) ────────")
    out.append("ALTER TABLE public.owner_account ADD COLUMN IF NOT EXISTS legacy_id text;")
    out.append("ALTER TABLE public.owner_account ADD COLUMN IF NOT EXISTS access_scope text;")
    out.append("ALTER TABLE public.owner_account ADD COLUMN IF NOT EXISTS notes text;")
    out.append("")

    # ── 2. UNIQUE індекси для ON CONFLICT ────────────────────────────────────
    out.append("-- ── 2. UNIQUE індекси — обовʼязкові для ON CONFLICT у INSERT'ах нижче ──")
    out.append("CREATE UNIQUE INDEX IF NOT EXISTS system_settings_tenant_sid_uidx ON public.system_settings (tenant_id, setting_id);")
    out.append("CREATE UNIQUE INDEX IF NOT EXISTS app_content_tenant_cid_uidx ON public.app_content (tenant_id, content_id);")
    out.append("CREATE UNIQUE INDEX IF NOT EXISTS distribution_template_tenant_tid_uidx ON public.distribution_template (tenant_id, tmpl_id);")
    out.append("CREATE UNIQUE INDEX IF NOT EXISTS notifications_tenant_nid_uidx ON public.notifications (tenant_id, notif_id);")
    out.append("CREATE UNIQUE INDEX IF NOT EXISTS clients_directory_tenant_cid_uidx ON public.clients_directory (tenant_id, cli_id);")
    out.append("CREATE UNIQUE INDEX IF NOT EXISTS staff_tenant_sid_uidx ON public.staff (tenant_id, staff_id);")
    out.append("CREATE UNIQUE INDEX IF NOT EXISTS owner_account_tenant_login_uidx ON public.owner_account (tenant_id, login);")
    out.append("")

    # ── 3. INSERT'и ──────────────────────────────────────────────────────────
    # Додаткові власники (крім primary) йдуть у staff як role='owner'
    extra_owner_staff = build_extra_owners_as_staff(owners)

    sections = [
        ('system_settings',       build_system_settings(settings)),
        ('app_content',           build_app_content(app_content)),
        ('distribution_template', build_distribution_template(templates)),
        ('notifications',         build_notifications(notifications)),
        ('clients_directory',     build_clients_directory(clients)),
        ('staff',                 build_staff(staff) + extra_owner_staff),
        ('owner_account',         build_owner_account(owners)),
    ]

    for name, stmts in sections:
        out.append(f"-- ── {name} ({len(stmts)} рядків) ─────────────────────────────────────")
        out.extend(stmts)
        out.append("")

    out.append("COMMIT;")
    out.append("")
    out.append("-- Оновити кеш PostgREST.")
    out.append("NOTIFY pgrst, 'reload schema';")

    with open(OUT, 'w', encoding='utf-8') as f:
        f.write('\n'.join(out))

    print(f'✅ Згенеровано: {OUT}')
    print(f'   Загалом ~{total} INSERT\'ів ({sum(len(s) for _, s in sections)} реально):')
    for name, stmts in sections:
        print(f'   {name:<25} {len(stmts):>4}')


if __name__ == '__main__':
    main()
