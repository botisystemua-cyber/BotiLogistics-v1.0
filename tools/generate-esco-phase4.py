#!/usr/bin/env python3
# ============================================================================
# generate-esco-phase4.py — payments + routes + dispatches + orders
# ============================================================================
#
# 61 payments + 8 routes + 2 dispatches + 1 order = 72 рядки.
#
# ЗМІНА СХЕМИ
#   ALTER TABLE routes ADD COLUMN IF NOT EXISTS extra_data jsonb
#   ALTER TABLE dispatches ALTER COLUMN route_id DROP NOT NULL
#   ALTER TABLE dispatches ALTER COLUMN vehicle_id DROP NOT NULL
#   (dispatches з xlsx мають порожні RTE_ID/AUTO_ID — без DROP NOT NULL
#   2 рядки не вставляться, що порушить принцип «нічого не загубити».)
#
# FK lookup'и:
#   payments.passenger_id  ← (SELECT id FROM passengers WHERE pax_id=…)
#   payments.calendar_id   ← (SELECT id FROM calendar WHERE cal_id=…)
#   payments.route_id      ← (SELECT id FROM routes WHERE rte_id=…) — у same TX
#   orders.client_id       ← (SELECT id FROM clients_directory WHERE cli_id=…)
#   dispatches.route_id    ← (SELECT id FROM routes WHERE rte_id=…) або NULL
#   dispatches.vehicle_id  ← (SELECT id FROM vehicles WHERE auto_id=…) або NULL
#
# Порядок INSERT:
#   1. routes (8) — щоб payments/dispatches могли резолвити route_id
#   2. payments (61)
#   3. dispatches (2)
#   4. orders (1)
# ============================================================================

import openpyxl
import os
import json
import re
from datetime import datetime, date

TENANT = 'esco'
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, 'sql', '2026-04-esco-phase4-finance-routes.sql')
XLSX_FIN = os.path.join(ROOT, 'PerenosTablu', 'Finance_crm_v2.xlsx')
XLSX_MARH = os.path.join(ROOT, 'PerenosTablu', 'Marhrut_crm_v6.xlsx')
XLSX_KLI = os.path.join(ROOT, 'PerenosTablu', 'Kliyentu_crm_v3.xlsx')


# ── helpers (compact) ────────────────────────────────────────────────────

def q(v):
    if v is None: return 'NULL'
    if isinstance(v, bool): return 'true' if v else 'false'
    if isinstance(v, (int, float)):
        if isinstance(v, float) and v.is_integer(): return str(int(v))
        return str(v)
    if isinstance(v, datetime): return "'" + v.strftime('%Y-%m-%d %H:%M:%S') + "'::timestamptz"
    if isinstance(v, date): return "'" + v.strftime('%Y-%m-%d') + "'::date"
    s = str(v).replace("'", "''")
    return "'" + s + "'"


def q_jsonb(obj):
    def _ser(v):
        if isinstance(v, (datetime, date)): return v.isoformat()
        return v
    cleaned = {k: _ser(v) for k, v in obj.items() if v is not None and v != ''}
    if not cleaned: return 'NULL'
    s = json.dumps(cleaned, ensure_ascii=False).replace("'", "''")
    return "'" + s + "'::jsonb"


def nn(v):
    if v is None: return None
    if isinstance(v, str):
        s = v.strip()
        if not s or s.lower() in ('none', 'null', '#n/a', '#value!'): return None
        return s
    return v


def to_num(v):
    v = nn(v)
    if v is None: return None
    if isinstance(v, (int, float)): return v
    s = re.sub(r'[^\d.\-]', '', str(v).replace(',', '.'))
    if not s or s in ('.', '-'): return None
    try:
        return float(s) if '.' in s else int(s)
    except ValueError:
        return None


def to_date_any(v):
    v = nn(v)
    if v is None: return None
    if isinstance(v, (datetime, date)): return v
    s = str(v).strip()
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S.%fZ', '%Y-%m-%dT%H:%M:%S',
                '%d.%m.%Y %H:%M', '%d.%m.%Y', '%Y-%m-%d'):
        try: return datetime.strptime(s, fmt)
        except ValueError: pass
    return None


def to_date_only(v):
    d = to_date_any(v)
    if d is None: return None
    return d.date() if isinstance(d, datetime) else d


def norm_phone(v):
    v = nn(v)
    if v is None: return None
    if isinstance(v, float) and v.is_integer(): v = str(int(v))
    s = re.sub(r'[\s()\-]', '', str(v))
    if re.match(r'^\d+\.0$', s): s = s[:-2]
    if s.startswith('+'): return s
    if re.match(r'^380\d{9}$', s): return '+' + s
    if re.match(r'^0\d{9}$', s): return '+38' + s
    if re.match(r'^\d{10,15}$', s): return '+' + s
    return s


def lookup(table, key_col, key_val):
    key_val = nn(key_val)
    if not key_val: return 'NULL'
    return (f"(SELECT id FROM public.{table} WHERE tenant_id={q(TENANT)} "
            f"AND {key_col}={q(key_val)} LIMIT 1)")


def read_sheet(path, name):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return []
    headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    out = []
    for r in rows[1:]:
        if all(v is None or v == '' for v in r): continue
        d = {}
        for i, h in enumerate(headers):
            d[h] = r[i] if i < len(r) else None
        out.append(d)
    wb.close()
    return out


# ── builders ────────────────────────────────────────────────────────────

def build_routes(rows_zurich, rows_geneva):
    sql = []
    for r in rows_zurich + rows_geneva:
        rte_id = nn(r.get('RTE_ID'))
        if not rte_id: continue

        # Розширені поля (xlsx 51 колонка vs БД ~28-48) → у extra_data.
        extra = {
            'source_sheet': nn(r.get('SOURCE_SHEET')),
            'created_at_xlsx': r.get('Дата створення'),
            'sender_phone': norm_phone(r.get('Телефон пасажира')),  # is_passenger==Пасажир
            'recipient_address': nn(r.get('Адреса отримувача')),
            'package_internal_no': nn(r.get('Внутрішній №')),
            'ttn_number': nn(r.get('Номер ТТН')),
            'package_description': nn(r.get('Опис посилки')),
            'package_kg': to_num(r.get('Кг посилки')),
            'amount': to_num(r.get('Сума')),
            'amount_currency': nn(r.get('Валюта')),
            'deposit': to_num(r.get('Завдаток')),
            'deposit_currency': nn(r.get('Валюта завдатку')),
            'payment_form': nn(r.get('Форма оплати')),
            'payment_status': nn(r.get('Статус оплати')),
            'debt': to_num(r.get('Борг')),
            'payment_note': nn(r.get('Примітка оплати')),
            'crm_status': nn(r.get('Статус CRM')),
            'tag': nn(r.get('Тег')),
            'driver_rating': to_num(r.get('Рейтинг водія')),
            'driver_comment': nn(r.get('Коментар водія')),
            'manager_rating': to_num(r.get('Рейтинг менеджера')),
            'manager_comment': nn(r.get('Коментар менеджера')),
            'sms_note': nn(r.get('Примітка СМС')),
            'cli_id': nn(r.get('CLI_ID')),
            'photo_url': nn(r.get('Фото посилки')),
            'archive_id': nn(r.get('ARCHIVE_ID')),
            'archived_by': nn(r.get('ARCHIVED_BY')),
            'archive_reason': nn(r.get('ARCHIVE_REASON')),
            'archive_date': r.get('DATE_ARCHIVE'),
        }

        sql.append(
            f"INSERT INTO public.routes ("
            f"tenant_id, rte_id, record_type, direction, pax_id_or_pkg_id, "
            f"route_date, timing, vehicle_name, driver_name, driver_phone, city, "
            f"seat_number, passenger_name, passenger_phone, "
            f"sender_name, recipient_name, recipient_phone, recipient_address, "
            f"departure_address, arrival_address, seats_count, baggage_weight, "
            f"internal_number, ttn_number, extra_data) VALUES ("
            f"{q(TENANT)}, {q(rte_id)}, {q(nn(r.get('Тип запису')))}, "
            f"{q(nn(r.get('Напрям')))}, {q(nn(r.get('PAX_ID / PKG_ID')))}, "
            f"{q(str(r.get('Дата рейсу')) if r.get('Дата рейсу') else None)}, "
            f"{q(nn(r.get('Таймінг')))}, {q(nn(r.get('Номер авто')))}, "
            f"{q(nn(r.get('Водій')))}, {q(norm_phone(r.get('Телефон водія')))}, "
            f"{q(nn(r.get('Місто')))}, {q(nn(r.get('Місце в авто')))}, "
            f"{q(nn(r.get('Піб пасажира')))}, {q(norm_phone(r.get('Телефон пасажира')))}, "
            f"{q(nn(r.get('Піб відправника')))}, {q(nn(r.get('Піб отримувача')))}, "
            f"{q(norm_phone(r.get('Телефон отримувача')))}, {q(nn(r.get('Адреса отримувача')))}, "
            f"{q(nn(r.get('Адреса відправки')))}, {q(nn(r.get('Адреса прибуття')))}, "
            f"{q(str(to_num(r.get('Кількість місць'))) if to_num(r.get('Кількість місць')) is not None else None)}, "
            f"{q(str(to_num(r.get('Вага багажу'))) if to_num(r.get('Вага багажу')) is not None else None)}, "
            f"{q(nn(r.get('Внутрішній №')))}, {q(nn(r.get('Номер ТТН')))}, "
            f"{q_jsonb(extra)}"
            f") ON CONFLICT (tenant_id, rte_id) WHERE tenant_id = 'esco' DO NOTHING;"
        )
    return sql


def build_payments(rows):
    sql = []
    for r in rows:
        pay_id = nn(r.get('PAY_ID'))
        if not pay_id: continue
        amount = to_num(r.get('Сума')) or 0
        ptype = nn(r.get('Тип платежу')) or 'unknown'
        currency = nn(r.get('Валюта')) or 'UAH'

        notes_parts = []
        if nn(r.get('Хто вніс')):
            notes_parts.append(f"Хто вніс: {nn(r.get('Хто вніс'))}")
        if nn(r.get('Примітка')):
            notes_parts.append(str(nn(r.get('Примітка'))))
        if to_num(r.get('Чайові')):
            notes_parts.append(f"Чайові: {to_num(r.get('Чайові'))}")
        if nn(r.get('ARCHIVED_BY')):
            notes_parts.append(f"Архівував: {nn(r.get('ARCHIVED_BY'))}")
        notes = ' | '.join(notes_parts) if notes_parts else None

        sql.append(
            f"INSERT INTO public.payments ("
            f"tenant_id, pay_id, created_at, created_by_role, "
            f"passenger_id, route_id, calendar_id, smart_id, "
            f"payment_type, amount, amount_currency, payment_form, payment_status, "
            f"debt_amount, debt_currency, due_date, notes) VALUES ("
            f"{q(TENANT)}, {q(pay_id)}, {q(to_date_any(r.get('Дата створення')))}, "
            f"{q(nn(r.get('Роль')))}, "
            f"{lookup('passengers', 'pax_id', r.get('PAX_ID'))}, "
            f"{lookup('routes', 'rte_id', r.get('RTE_ID'))}, "
            f"{lookup('calendar', 'cal_id', r.get('CAL_ID'))}, "
            f"{q(nn(r.get('Ід_смарт')))}, {q(ptype)}, {q(amount)}, {q(currency)}, "
            f"{q(nn(r.get('Форма оплати')))}, {q(nn(r.get('Статус платежу')))}, "
            f"{q(to_num(r.get('Борг сума')))}, {q(nn(r.get('Борг валюта')) or 'UAH')}, "
            f"{q(to_date_only(r.get('Дата погашення')))}, {q(notes)}"
            f") ON CONFLICT (tenant_id, pay_id) WHERE tenant_id = 'esco' DO NOTHING;"
        )
    return sql


def build_dispatches(rows_zurich, rows_geneva):
    sql = []
    for r in rows_zurich + rows_geneva:
        d_id = nn(r.get('DISPATCH_ID'))
        if not d_id: continue
        rte_id = nn(r.get('RTE_ID'))
        auto_id = nn(r.get('AUTO_ID'))
        route_date = to_date_only(r.get('Дата рейсу'))

        notes_parts = []
        if nn(r.get('Водій')):
            notes_parts.append(f"Водій: {nn(r.get('Водій'))}")
        if nn(r.get('Примітка')):
            notes_parts.append(str(nn(r.get('Примітка'))))
        if not rte_id:
            notes_parts.append('RTE_ID(raw)=порожнє у джерелі')
        if not auto_id:
            notes_parts.append('AUTO_ID(raw)=порожнє у джерелі')
        notes = ' | '.join(notes_parts) if notes_parts else None

        sql.append(
            f"INSERT INTO public.dispatches ("
            f"tenant_id, dispatch_id, created_at, route_id, route_date, vehicle_id, "
            f"sender_name, sender_phone, registrar_phone, "
            f"recipient_name, recipient_phone, recipient_address, internal_number, "
            f"weight_kg, package_description, photo_url, "
            f"amount, amount_currency, deposit, deposit_currency, "
            f"payment_form, payment_status, debt, status, notes) VALUES ("
            f"{q(TENANT)}, {q(d_id)}, {q(to_date_any(r.get('Дата створення')))}, "
            f"{lookup('routes', 'rte_id', rte_id)}, {q(route_date)}, "
            f"{lookup('vehicles', 'auto_id', auto_id)}, "
            f"{q(nn(r.get('Піб відправника')))}, "
            f"{q(norm_phone(r.get('Телефон відправника')))}, "
            f"{q(norm_phone(r.get('Телефон відправника')))}, "
            f"{q(nn(r.get('Піб отримувача')))}, "
            f"{q(norm_phone(r.get('Телефон отримувача')))}, "
            f"{q(nn(r.get('Адреса отримувача')))}, {q(nn(r.get('Внутрішній №')))}, "
            f"{q(to_num(r.get('Вага')))}, {q(nn(r.get('Опис посилки')))}, "
            f"{q(nn(r.get('Фото посилки')))}, "
            f"{q(to_num(r.get('Сума')))}, {q(nn(r.get('Валюта')) or 'UAH')}, "
            f"{q(to_num(r.get('Завдаток')))}, {q(nn(r.get('Валюта завдатку')) or 'UAH')}, "
            f"{q(nn(r.get('Форма оплати')))}, {q(nn(r.get('Статус оплати')) or 'pending')}, "
            f"{q(to_num(r.get('Борг')))}, {q(nn(r.get('Статус')) or 'pending')}, "
            f"{q(notes)}"
            f") ON CONFLICT (tenant_id, dispatch_id) WHERE tenant_id = 'esco' DO NOTHING;"
        )
    return sql


def build_orders(rows):
    sql = []
    for r in rows:
        order_id = nn(r.get('ORDER_ID'))
        if not order_id: continue
        cli_id = nn(r.get('CLIENT_ID'))
        if not cli_id:
            continue  # client_id NOT NULL, skip
        sql.append(
            f"INSERT INTO public.orders ("
            f"tenant_id, order_id, client_id, created_at, direction, "
            f"sender_address, recipient_address, recipient_phone, "
            f"weight_kg, description, photo_url, "
            f"price, price_currency, payment_status, package_status, "
            f"delivery_date, client_notes, manager_notes) VALUES ("
            f"{q(TENANT)}, {q(order_id)}, "
            f"{lookup('clients_directory', 'cli_id', cli_id)}, "
            f"{q(to_date_any(r.get('Дата створення')))}, "
            f"{q(nn(r.get('Напрям')) or 'UA → EU')}, "
            f"{q(nn(r.get('Адреса відправника')) or '(не вказано)')}, "
            f"{q(nn(r.get('Адреса отримувача')) or '(не вказано)')}, "
            f"{q(norm_phone(r.get('Телефон отримувача')))}, "
            f"{q(to_num(r.get('Вага')))}, {q(nn(r.get('Опис')))}, "
            f"{q(nn(r.get('Фото')))}, "
            f"{q(to_num(r.get('Ціна')))}, {q(nn(r.get('Валюта')) or 'UAH')}, "
            f"{q(nn(r.get('Статус оплати')) or 'pending')}, "
            f"{q(nn(r.get('Статус посилки')) or 'pending')}, "
            f"{q(to_date_only(r.get('Дата доставки')))}, "
            f"{q(nn(r.get('Примітка клієнта')))}, {q(nn(r.get('Примітка менеджера')))}"
            f") ON CONFLICT (tenant_id, order_id) WHERE tenant_id = 'esco' DO NOTHING;"
        )
    return sql


def main():
    routes_z = read_sheet(XLSX_MARH, 'Маршрут_Цюріх')
    routes_g = read_sheet(XLSX_MARH, 'Маршрут_Женева')
    disp_z = read_sheet(XLSX_MARH, 'Відправка_Цюріх')
    disp_g = read_sheet(XLSX_MARH, 'Відправка_Женева')
    payments = read_sheet(XLSX_FIN, 'Платежі')
    orders = read_sheet(XLSX_KLI, 'Замовлення')

    routes_sql = build_routes(routes_z, routes_g)
    payments_sql = build_payments(payments)
    dispatches_sql = build_dispatches(disp_z, disp_g)
    orders_sql = build_orders(orders)

    out = []
    out.append("-- =========================================================================")
    out.append("-- ESCO MIGRATION — PHASE 4: PAYMENTS + ROUTES + DISPATCHES + ORDERS")
    out.append("-- =========================================================================")
    out.append(f"-- routes:     {len(routes_sql)}")
    out.append(f"-- payments:   {len(payments_sql)}")
    out.append(f"-- dispatches: {len(dispatches_sql)}")
    out.append(f"-- orders:     {len(orders_sql)}")
    out.append("--")
    out.append("-- ВАЖЛИВО:")
    out.append("--   ALTER dispatches DROP NOT NULL для route_id, vehicle_id —")
    out.append("--   у xlsx обидва Відправки мають порожні RTE_ID/AUTO_ID,")
    out.append("--   без DROP NOT NULL 2 рядки не зайдуть → втратимо.")
    out.append("--")
    out.append("-- Порядок: routes → payments/dispatches/orders (FK lookup'и).")
    out.append("-- =========================================================================")
    out.append("")
    out.append("BEGIN;")
    out.append("")
    out.append("-- ── Schema changes ──────────────────────────────────────────────")
    out.append("ALTER TABLE public.routes ADD COLUMN IF NOT EXISTS extra_data jsonb;")
    out.append("ALTER TABLE public.dispatches ALTER COLUMN route_id DROP NOT NULL;")
    out.append("ALTER TABLE public.dispatches ALTER COLUMN vehicle_id DROP NOT NULL;")
    out.append("ALTER TABLE public.dispatches ALTER COLUMN route_date DROP NOT NULL;")
    out.append("")
    out.append("-- ── UNIQUE indexes (PARTIAL — лише для tenant='esco', щоб не зачіпати ──")
    out.append("--    дублі інших тенантів типу express_sv_travel) ─────────────────")
    out.append("CREATE UNIQUE INDEX IF NOT EXISTS routes_esco_rte_uidx ON public.routes (tenant_id, rte_id) WHERE tenant_id = 'esco';")
    out.append("CREATE UNIQUE INDEX IF NOT EXISTS payments_esco_pay_uidx ON public.payments (tenant_id, pay_id) WHERE tenant_id = 'esco';")
    out.append("CREATE UNIQUE INDEX IF NOT EXISTS dispatches_esco_disp_uidx ON public.dispatches (tenant_id, dispatch_id) WHERE tenant_id = 'esco';")
    out.append("CREATE UNIQUE INDEX IF NOT EXISTS orders_esco_order_uidx ON public.orders (tenant_id, order_id) WHERE tenant_id = 'esco';")
    out.append("")
    out.append(f"-- ── routes ({len(routes_sql)}) ───────────────────────────────────")
    out.extend(routes_sql)
    out.append("")
    out.append(f"-- ── payments ({len(payments_sql)}) ───────────────────────────────")
    out.extend(payments_sql)
    out.append("")
    out.append(f"-- ── dispatches ({len(dispatches_sql)}) ───────────────────────────")
    out.extend(dispatches_sql)
    out.append("")
    out.append(f"-- ── orders ({len(orders_sql)}) ───────────────────────────────────")
    out.extend(orders_sql)
    out.append("")
    out.append("COMMIT;")
    out.append("")
    out.append("NOTIFY pgrst, 'reload schema';")

    with open(OUT, 'w', encoding='utf-8') as f:
        f.write('\n'.join(out))

    total = len(routes_sql) + len(payments_sql) + len(dispatches_sql) + len(orders_sql)
    print(f'✅ {OUT}')
    print(f'   routes:     {len(routes_sql)}')
    print(f'   payments:   {len(payments_sql)}')
    print(f'   dispatches: {len(dispatches_sql)}')
    print(f'   orders:     {len(orders_sql)}')
    print(f'   ──────────────')
    print(f'   total:      {total}')


if __name__ == '__main__':
    main()
