#!/usr/bin/env python3
# ============================================================================
# generate-esco-phase5.py — Phase 5: archive_packages + audit/change/access logs
# ============================================================================
#
# ОБСЯГИ
#   archive_packages:  9   (Archive_crm/Посилки)
#   change_logs:    1181   (Archive_crm/Логи)         — основний логовий обсяг
#   audit_logs:        1   (Marhrut_crm/Логи водіїв)
#   access_logs:      84   (Config_crm/Лог доступів)
#   ───────────────────────
#   total:          1275
#
# ЗМІНИ СХЕМИ
#   ALTER archive_packages ADD COLUMN extra_data jsonb
#     (xlsx 46 полів vs БД 29 → зайве у extra_data)
#   ALTER change_logs ADD COLUMN changed_by_name text
#     (changed_by — uuid, а у xlsx ім'я текстом)
#
# ОСОБЛИВОСТІ
#   ip_address у БД — INET; парсимо тільки валідні IPv4, інакше NULL.
#   audit_logs не має log_id → один INSERT без ON CONFLICT (1 рядок).
#   Партіал-індекси WHERE tenant_id='esco' (як у Phase 4).
# ============================================================================

import openpyxl
import os
import json
import re
from datetime import datetime, date

TENANT = 'esco'
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, 'sql', '2026-04-esco-phase5-archive-logs.sql')
XLSX_ARCH = os.path.join(ROOT, 'PerenosTablu', 'Archive_crm_v3.xlsx')
XLSX_MARH = os.path.join(ROOT, 'PerenosTablu', 'Marhrut_crm_v6.xlsx')
XLSX_CONF = os.path.join(ROOT, 'PerenosTablu', 'Config_crm_v2.xlsx')


def q(v):
    if v is None: return 'NULL'
    if isinstance(v, bool): return 'true' if v else 'false'
    if isinstance(v, (int, float)):
        if isinstance(v, float) and v.is_integer(): return str(int(v))
        return str(v)
    if isinstance(v, datetime): return "'" + v.strftime('%Y-%m-%d %H:%M:%S') + "'::timestamptz"
    if isinstance(v, date): return "'" + v.strftime('%Y-%m-%d') + "'::date"
    s = str(v).replace("'", "''")
    return "'" + s + "'"


def q_jsonb(obj):
    def _ser(v):
        if isinstance(v, (datetime, date)): return v.isoformat()
        return v
    cleaned = {k: _ser(v) for k, v in obj.items() if v is not None and v != ''}
    if not cleaned: return 'NULL'
    s = json.dumps(cleaned, ensure_ascii=False).replace("'", "''")
    return "'" + s + "'::jsonb"


def q_inet(v):
    """ip_address як INET. Тільки валідні IPv4, інакше NULL."""
    v = nn(v)
    if v is None:
        return 'NULL'
    s = str(v).strip()
    # IPv4 матч
    if re.match(r'^(\d{1,3}\.){3}\d{1,3}$', s):
        # Перевірка що кожен октет 0-255
        parts = s.split('.')
        if all(0 <= int(p) <= 255 for p in parts):
            return f"'{s}'::inet"
    return 'NULL'


def nn(v):
    if v is None: return None
    if isinstance(v, str):
        s = v.strip()
        if not s or s.lower() in ('none', 'null', '#n/a', '#value!'): return None
        return s
    return v


def to_num(v):
    v = nn(v)
    if v is None: return None
    if isinstance(v, (int, float)): return v
    s = re.sub(r'[^\d.\-]', '', str(v).replace(',', '.'))
    if not s or s in ('.', '-'): return None
    try:
        return float(s) if '.' in s else int(s)
    except ValueError:
        return None


def to_bool(v):
    v = nn(v)
    if v is None: return None
    if isinstance(v, bool): return v
    s = str(v).strip().lower()
    if s in ('так', 'yes', 'true', '1', 'y'): return True
    if s in ('ні', 'no', 'false', '0', 'n'): return False
    return None


def to_date_any(v):
    v = nn(v)
    if v is None: return None
    if isinstance(v, (datetime, date)): return v
    s = str(v).strip()
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S.%fZ', '%Y-%m-%dT%H:%M:%S',
                '%d.%m.%Y %H:%M', '%d.%m.%Y', '%Y-%m-%d', '%H:%M:%S', '%H:%M'):
        try: return datetime.strptime(s, fmt)
        except ValueError: pass
    return None


def norm_phone(v):
    v = nn(v)
    if v is None: return None
    if isinstance(v, float) and v.is_integer(): v = str(int(v))
    s = re.sub(r'[\s()\-]', '', str(v))
    if re.match(r'^\d+\.0$', s): s = s[:-2]
    if s.startswith('+'): return s
    if re.match(r'^380\d{9}$', s): return '+' + s
    if re.match(r'^0\d{9}$', s): return '+38' + s
    if re.match(r'^\d{10,15}$', s): return '+' + s
    return s


def read_sheet(path, name):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return []
    headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    out = []
    for r in rows[1:]:
        if all(v is None or v == '' for v in r): continue
        d = {}
        for i, h in enumerate(headers):
            d[h] = r[i] if i < len(r) else None
        out.append(d)
    wb.close()
    return out


# ── builders ────────────────────────────────────────────────────────────

def build_archive_packages(rows):
    sql = []
    for r in rows:
        archive_id = nn(r.get('ARCHIVE_ID'))
        if not archive_id: continue
        action_type = nn(r.get('Тип дії')) or 'archive'
        extra = {
            'cli_id': nn(r.get('CLI_ID')),
            'rte_id': nn(r.get('RTE_ID')),
            'smart_id': nn(r.get('Ід_смарт')),
            'direction': nn(r.get('Напрям')),
            'registrar_phone': norm_phone(r.get('Телефон реєстратора')),
            'recipient_phone': norm_phone(r.get('Телефон отримувача')),
            'evaluated_value': to_num(r.get('Оціночна вартість')),
            'cod_amount': to_num(r.get('Сума НП')),
            'cod_currency': nn(r.get('Валюта НП')),
            'cod_status': nn(r.get('Статус НП')),
            'deposit_currency': nn(r.get('Валюта завдатку')),
            'payment_form': nn(r.get('Форма оплати')),
            'departure_date': r.get('Дата відправки'),
            'departure_time': nn(r.get('Таймінг')),
            'auto_id': nn(r.get('AUTO_ID')),
            'vehicle_number': nn(r.get('Номер авто')),
            'status_at_archive': nn(r.get('Статус на момент')),
            'crm_status': nn(r.get('Статус CRM')),
            'driver_rating': to_num(r.get('Рейтинг водія')),
            'driver_comment': nn(r.get('Коментар водія')),
            'manager_rating': to_num(r.get('Рейтинг менеджера')),
            'manager_comment': nn(r.get('Коментар менеджера')),
            'archived_by_name': nn(r.get('ARCHIVED_BY')),
            'restored_by_name': nn(r.get('Відновив')),
        }
        sql.append(
            f"INSERT INTO public.archive_packages ("
            f"tenant_id, archive_id, action_type, archive_date, archived_by, archive_reason, "
            f"source_table, source_sheet, pkg_id, sender_name, recipient_name, recipient_address, "
            f"internal_number, ttn_number, description, weight_kg, amount, amount_currency, "
            f"deposit, debt, payment_status, package_status, "
            f"restored, restored_date, restoration_reason, extra_data) VALUES ("
            f"{q(TENANT)}, {q(archive_id)}, {q(action_type)}, "
            f"{q(to_date_any(r.get('DATE_ARCHIVE')))}, NULL, "
            f"{q(nn(r.get('ARCHIVE_REASON')))}, "
            f"{q(nn(r.get('SOURCE_TABLE')) or 'packages')}, {q(nn(r.get('SOURCE_SHEET')))}, "
            f"{q(nn(r.get('PKG_ID')))}, {q(nn(r.get('Піб відправника')))}, "
            f"{q(nn(r.get('Піб отримувача')))}, {q(nn(r.get('Адреса отримувача')))}, "
            f"{q(nn(r.get('Внутрішній №')))}, {q(nn(r.get('Номер ТТН')))}, "
            f"{q(nn(r.get('Опис посилки')))}, {q(to_num(r.get('Кг')))}, "
            f"{q(to_num(r.get('Сума')))}, {q(nn(r.get('Валюта оплати')))}, "
            f"{q(to_num(r.get('Завдаток')))}, {q(to_num(r.get('Борг')))}, "
            f"{q(nn(r.get('Статус оплати')))}, {q(nn(r.get('Статус CRM')))}, "
            f"{q(bool(to_num(r.get('Відновлено'))))}, "
            f"{q(to_date_any(r.get('Дата відновлення')))}, "
            f"{q(nn(r.get('Причина відновлення')))}, "
            f"{q_jsonb(extra)}"
            f") ON CONFLICT (tenant_id, archive_id) WHERE tenant_id = 'esco' DO NOTHING;"
        )
    return sql


def build_change_logs(rows):
    sql = []
    for r in rows:
        log_id = nn(r.get('LOG_ID'))
        if not log_id: continue
        action = nn(r.get('Дія')) or 'unknown'
        table_name = nn(r.get('Таблиця')) or 'unknown'
        sql.append(
            f"INSERT INTO public.change_logs ("
            f"tenant_id, log_id, change_date, changed_by_name, changed_by_role, "
            f"action, table_name, sheet_name, record_id, field_name, "
            f"old_value, new_value, ip_address, device_info, "
            f"verified_by_owner, verified_date, is_suspicious, notes) VALUES ("
            f"{q(TENANT)}, {q(log_id)}, {q(to_date_any(r.get('Дата і час')))}, "
            f"{q(nn(r.get('Хто')))}, {q(nn(r.get('Роль')))}, "
            f"{q(action)}, {q(table_name)}, {q(nn(r.get('Аркуш')))}, "
            f"{q(nn(r.get('ID запису')))}, {q(nn(r.get('Поле')))}, "
            f"{q(nn(r.get('Значення БУЛО')))}, {q(nn(r.get('Значення СТАЛО')))}, "
            f"{q_inet(r.get('IP адреса'))}, {q(nn(r.get('Пристрій')))}, "
            f"{q(to_bool(r.get('Підтверджено власником')))}, "
            f"{q(to_date_any(r.get('Дата підтвердження')))}, "
            f"{q(to_bool(r.get('Підозріло')))}, {q(nn(r.get('Примітка')))}"
            f") ON CONFLICT (tenant_id, log_id) WHERE tenant_id = 'esco' DO NOTHING;"
        )
    return sql


def build_audit_logs(rows):
    sql = []
    for r in rows:
        # Marhrut/Логи водіїв: Дата, Час, Водій, Маршрут, ID запису, Тип, Статус, Причина, Телефон
        action = nn(r.get('Тип')) or 'unknown'
        table_name = nn(r.get('Маршрут')) or 'unknown'
        # Combine Дата + Час → datetime
        dt = to_date_any(r.get('Дата'))
        time_part = r.get('Час')
        if dt and time_part:
            try:
                if hasattr(time_part, 'hour'):
                    dt = dt.replace(hour=time_part.hour, minute=time_part.minute, second=getattr(time_part, 'second', 0))
            except Exception:
                pass

        new_vals = {
            'driver': nn(r.get('Водій')),
            'status': nn(r.get('Статус')),
            'reason': nn(r.get('Причина')),
            'phone': norm_phone(r.get('Телефон')),
        }

        sql.append(
            f"INSERT INTO public.audit_logs ("
            f"tenant_id, table_name, action, record_id, new_values, created_at) VALUES ("
            f"{q(TENANT)}, {q(table_name)}, {q(action)}, "
            f"{q(nn(r.get('ID запису')))}, {q_jsonb(new_vals)}, {q(dt)}"
            f");"
        )
    return sql


def build_access_logs(rows):
    sql = []
    for r in rows:
        log_id = nn(r.get('LOG_ID'))
        if not log_id: continue
        action = nn(r.get('Дія')) or 'unknown'
        sql.append(
            f"INSERT INTO public.access_logs ("
            f"tenant_id, log_id, user_name, user_role, action, "
            f"table_accessed, sheet_accessed, ip_address, device_info, "
            f"access_status, failure_reason, created_at) VALUES ("
            f"{q(TENANT)}, {q(log_id)}, {q(nn(r.get('Піб')))}, {q(nn(r.get('Роль')))}, "
            f"{q(action)}, {q(nn(r.get('Таблиця')))}, {q(nn(r.get('Аркуш')))}, "
            f"{q_inet(r.get('IP адреса'))}, {q(nn(r.get('Пристрій')))}, "
            f"{q(nn(r.get('Статус')) or 'success')}, {q(nn(r.get('Примітка')))}, "
            f"{q(to_date_any(r.get('Дата і час')))}"
            f") ON CONFLICT (tenant_id, log_id) WHERE tenant_id = 'esco' DO NOTHING;"
        )
    return sql


def main():
    arch_pkg = read_sheet(XLSX_ARCH, 'Посилки')
    arch_logs = read_sheet(XLSX_ARCH, 'Логи')
    driver_logs = read_sheet(XLSX_MARH, 'Логи водіїв')
    access = read_sheet(XLSX_CONF, 'Лог доступів')

    arch_pkg_sql = build_archive_packages(arch_pkg)
    change_sql = build_change_logs(arch_logs)
    audit_sql = build_audit_logs(driver_logs)
    access_sql = build_access_logs(access)

    out = []
    out.append("-- =========================================================================")
    out.append("-- ESCO MIGRATION — PHASE 5: ARCHIVE_PACKAGES + LOGS")
    out.append("-- =========================================================================")
    out.append(f"-- archive_packages: {len(arch_pkg_sql)}")
    out.append(f"-- change_logs:      {len(change_sql)}  (Архів/Логи)")
    out.append(f"-- audit_logs:       {len(audit_sql)}  (Marhrut/Логи водіїв)")
    out.append(f"-- access_logs:      {len(access_sql)} (Config/Лог доступів)")
    out.append(f"-- ────────────────────────────────")
    out.append(f"-- total:            {len(arch_pkg_sql)+len(change_sql)+len(audit_sql)+len(access_sql)}")
    out.append("-- =========================================================================")
    out.append("")
    out.append("BEGIN;")
    out.append("")
    out.append("-- ── Schema changes ──────────────────────────────────────────────")
    out.append("ALTER TABLE public.archive_packages ADD COLUMN IF NOT EXISTS extra_data jsonb;")
    out.append("ALTER TABLE public.change_logs ADD COLUMN IF NOT EXISTS changed_by_name text;")
    out.append("")
    out.append("-- ── Partial UNIQUE indexes (WHERE tenant='esco') ────────────────")
    out.append("CREATE UNIQUE INDEX IF NOT EXISTS archive_packages_esco_arc_uidx ON public.archive_packages (tenant_id, archive_id) WHERE tenant_id = 'esco';")
    out.append("CREATE UNIQUE INDEX IF NOT EXISTS change_logs_esco_log_uidx ON public.change_logs (tenant_id, log_id) WHERE tenant_id = 'esco';")
    out.append("CREATE UNIQUE INDEX IF NOT EXISTS access_logs_esco_log_uidx ON public.access_logs (tenant_id, log_id) WHERE tenant_id = 'esco';")
    out.append("")
    out.append(f"-- ── archive_packages ({len(arch_pkg_sql)}) ───────────────────────")
    out.extend(arch_pkg_sql)
    out.append("")
    out.append(f"-- ── change_logs ({len(change_sql)}) — основний обсяг логів ──────")
    out.extend(change_sql)
    out.append("")
    out.append(f"-- ── audit_logs ({len(audit_sql)}) ──────────────────────────────")
    out.extend(audit_sql)
    out.append("")
    out.append(f"-- ── access_logs ({len(access_sql)}) ────────────────────────────")
    out.extend(access_sql)
    out.append("")
    out.append("COMMIT;")
    out.append("")
    out.append("NOTIFY pgrst, 'reload schema';")

    with open(OUT, 'w', encoding='utf-8') as f:
        f.write('\n'.join(out))

    total = len(arch_pkg_sql) + len(change_sql) + len(audit_sql) + len(access_sql)
    print(f'✅ {OUT}')
    print(f'   archive_packages: {len(arch_pkg_sql)}')
    print(f'   change_logs:      {len(change_sql)}')
    print(f'   audit_logs:       {len(audit_sql)}')
    print(f'   access_logs:      {len(access_sql)}')
    print(f'   ──────────────────')
    print(f'   total:            {total}')


if __name__ == '__main__':
    main()
