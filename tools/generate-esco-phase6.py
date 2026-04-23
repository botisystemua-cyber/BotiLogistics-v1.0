#!/usr/bin/env python3
# ============================================================================
# generate-esco-phase6.py — Phase 6: messages + route_access + reviews +
#                          client_ratings + client_app_access (FINAL)
# ============================================================================
#
# ОБСЯГИ
#   messages:           4   (Kliyentu/Чат)
#   route_access:       3   (Config/Маршрути_доступ)
#   client_app_access:  3   (Config/Клієнти_доступ — без password_hash)
#   reviews:            1   (Kliyentu/Відгуки клієнтів)
#   client_ratings:     1   (Kliyentu/Рейтинг клієнтів)
#   ────────────────────────
#   total:             12
#
# ЗМІНИ СХЕМИ
#   ALTER route_access ALTER COLUMN staff_id, route_id DROP NOT NULL
#   ALTER reviews          ADD COLUMN extra_data jsonb
#   ALTER client_ratings   ADD COLUMN extra_data jsonb
# ============================================================================

import openpyxl, os, json, re
from datetime import datetime, date

TENANT = 'esco'
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, 'sql', '2026-04-esco-phase6-misc.sql')
XLSX_KLI = os.path.join(ROOT, 'PerenosTablu', 'Kliyentu_crm_v3.xlsx')
XLSX_CONF = os.path.join(ROOT, 'PerenosTablu', 'Config_crm_v2.xlsx')


def q(v):
    if v is None: return 'NULL'
    if isinstance(v, bool): return 'true' if v else 'false'
    if isinstance(v, (int, float)):
        if isinstance(v, float) and v.is_integer(): return str(int(v))
        return str(v)
    if isinstance(v, datetime): return "'" + v.strftime('%Y-%m-%d %H:%M:%S') + "'::timestamptz"
    if isinstance(v, date): return "'" + v.strftime('%Y-%m-%d') + "'::date"
    s = str(v).replace("'", "''")
    return "'" + s + "'"


def q_jsonb(obj):
    def _ser(v):
        if isinstance(v, (datetime, date)): return v.isoformat()
        return v
    cleaned = {k: _ser(v) for k, v in obj.items() if v is not None and v != ''}
    if not cleaned: return 'NULL'
    s = json.dumps(cleaned, ensure_ascii=False).replace("'", "''")
    return "'" + s + "'::jsonb"


def nn(v):
    if v is None: return None
    if isinstance(v, str):
        s = v.strip()
        if not s or s.lower() in ('none', 'null', '#n/a', '#value!'): return None
        return s
    return v


def to_num(v):
    v = nn(v)
    if v is None: return None
    if isinstance(v, (int, float)): return v
    s = re.sub(r'[^\d.\-]', '', str(v).replace(',', '.'))
    if not s or s in ('.', '-'): return None
    try: return float(s) if '.' in s else int(s)
    except ValueError: return None


def to_bool(v):
    v = nn(v)
    if v is None: return None
    if isinstance(v, bool): return v
    s = str(v).strip().lower()
    if s in ('так', 'yes', 'true', '1', 'y'): return True
    if s in ('ні', 'no', 'false', '0', 'n'): return False
    return None


def to_rating(v, max_val=5):
    n = to_num(v)
    if n is None or n == 0: return None
    if 0 < n <= max_val: return int(n) if int(n) == n else n
    return None  # >5 — silently NULL


def to_date_any(v):
    v = nn(v)
    if v is None: return None
    if isinstance(v, (datetime, date)): return v
    s = str(v).strip()
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S.%fZ', '%Y-%m-%dT%H:%M:%S',
                '%d.%m.%Y %H:%M', '%d.%m.%Y', '%Y-%m-%d'):
        try: return datetime.strptime(s, fmt)
        except ValueError: pass
    return None


def to_date_only(v):
    d = to_date_any(v)
    if d is None: return None
    return d.date() if isinstance(d, datetime) else d


def norm_phone(v):
    v = nn(v)
    if v is None: return None
    if isinstance(v, float) and v.is_integer(): v = str(int(v))
    s = re.sub(r'[\s()\-]', '', str(v))
    if re.match(r'^\d+\.0$', s): s = s[:-2]
    if s.startswith('+'): return s
    if re.match(r'^380\d{9}$', s): return '+' + s
    if re.match(r'^0\d{9}$', s): return '+38' + s
    if re.match(r'^\d{10,15}$', s): return '+' + s
    return s


def lookup(table, key_col, key_val):
    key_val = nn(key_val)
    if not key_val: return 'NULL'
    return (f"(SELECT id FROM public.{table} WHERE tenant_id={q(TENANT)} "
            f"AND {key_col}={q(key_val)} LIMIT 1)")


def read_sheet(path, name):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if name not in wb.sheetnames:
        wb.close(); return []
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close(); return []
    headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    out = []
    for r in rows[1:]:
        if all(v is None or v == '' for v in r): continue
        d = {}
        for i, h in enumerate(headers):
            d[h] = r[i] if i < len(r) else None
        out.append(d)
    wb.close()
    return out


# ── builders ────────────────────────────────────────────────────────────

def build_messages(rows):
    sql = []
    for r in rows:
        msg_id = nn(r.get('MESSAGE_ID'))
        if not msg_id: continue
        cli_id = nn(r.get('CLIENT_ID'))
        if not cli_id: continue  # client_id NOT NULL
        text = nn(r.get('Текст повідомлення')) or '(порожнє)'
        order_lookup = lookup('orders', 'order_id', r.get('ORDER_ID'))
        sender_name_key = "Ім'я відправника"
        sender_name_val = nn(r.get(sender_name_key))
        sql.append(
            f"INSERT INTO public.messages ("
            f"tenant_id, message_id, client_id, created_at, "
            f"sender_role, sender_name, message_text, is_read, order_id) VALUES ("
            f"{q(TENANT)}, {q(msg_id)}, "
            f"{lookup('clients_directory', 'cli_id', cli_id)}, "
            f"{q(to_date_any(r.get('Дата і час')))}, "
            f"{q(nn(r.get('Роль відправника')))}, "
            f"{q(sender_name_val)}, "
            f"{q(text)}, "
            f"{q(to_bool(r.get('Прочитано')))}, "
            f"{order_lookup}"
            f") ON CONFLICT (tenant_id, message_id) WHERE tenant_id = 'esco' DO NOTHING;"
        )
    return sql


def build_route_access(rows):
    sql = []
    for r in rows:
        access_id = nn(r.get('ACCESS_ID'))
        if not access_id: continue
        notes_parts = []
        if nn(r.get('Маршрут')):
            notes_parts.append(f"Маршрут: {nn(r.get('Маршрут'))}")
        if nn(r.get('Хто надав')):
            notes_parts.append(f"Хто надав: {nn(r.get('Хто надав'))}")
        if nn(r.get('Примітка')):
            notes_parts.append(str(nn(r.get('Примітка'))))
        notes = ' | '.join(notes_parts) if notes_parts else None

        sql.append(
            f"INSERT INTO public.route_access ("
            f"tenant_id, access_id, staff_id, staff_name, staff_role, "
            f"route_id, access_from_date, access_to_date, access_level, "
            f"granted_date, access_status, notes) VALUES ("
            f"{q(TENANT)}, {q(access_id)}, "
            f"{lookup('staff', 'staff_id', r.get('STAFF_ID'))}, "
            f"{q(nn(r.get('Піб персоналу')))}, {q(nn(r.get('Роль')))}, "
            f"{lookup('routes', 'rte_id', r.get('RTE_ID'))}, "
            f"{q(to_date_only(r.get('Дата від')))}, {q(to_date_only(r.get('Дата до')))}, "
            f"{q(nn(r.get('Рівень доступу')))}, "
            f"{q(to_date_any(r.get('Дата надання')))}, "
            f"{q(nn(r.get('Статус')) or 'active')}, {q(notes)}"
            f") ON CONFLICT (tenant_id, access_id) WHERE tenant_id = 'esco' DO NOTHING;"
        )
    return sql


def build_client_app_access(rows):
    sql = []
    for r in rows:
        access_id = nn(r.get('ACCESS_ID'))
        if not access_id: continue
        cli_id = nn(r.get('CLI_ID'))
        if not cli_id: continue  # client_id NOT NULL
        sql.append(
            f"INSERT INTO public.client_app_access ("
            f"tenant_id, access_id, client_id, phone, email, login, "
            f"app_status, registered_date, last_login, last_device, "
            f"is_blocked, block_reason, notes) VALUES ("
            f"{q(TENANT)}, {q(access_id)}, "
            f"{lookup('clients_directory', 'cli_id', cli_id)}, "
            f"{q(norm_phone(r.get('Телефон')))}, {q(nn(r.get('EMAIL')))}, "
            f"{q(nn(r.get('Логін')))}, "
            f"{q(nn(r.get('Статус апки')) or 'active')}, "
            f"{q(to_date_any(r.get('Дата реєстрації')))}, "
            f"{q(to_date_any(r.get('Остання активність')))}, "
            f"{q(nn(r.get('Пристрій')))}, "
            f"{q(to_bool(r.get('Заблоковано')))}, "
            f"{q(nn(r.get('Причина блокування')))}, {q(nn(r.get('Примітка')))}"
            f") ON CONFLICT (tenant_id, access_id) WHERE tenant_id = 'esco' DO NOTHING;"
        )
    return sql


def build_reviews(rows):
    sql = []
    for r in rows:
        rid = nn(r.get('REVIEW_ID'))
        if not rid: continue
        cli_id = nn(r.get('CLIENT_ID'))
        if not cli_id: continue  # client_id NOT NULL
        extra = {
            'smart_id': nn(r.get('Ід_смарт/CRM')),
            'route_date': r.get('Дата рейсу'),
            'direction': nn(r.get('Напрям')),
            'vehicle_number': nn(r.get('Номер авто')),
            'driver_name': nn(r.get('Водій')),
            'processed_flag': nn(r.get('Опрацьовано')),
            'processed_by_name': nn(r.get('Хто опрацював')),
        }
        sql.append(
            f"INSERT INTO public.reviews ("
            f"tenant_id, review_id, review_date, review_status, "
            f"route_id, passenger_id, client_id, client_phone, client_name, "
            f"record_type, driver_rating, driver_comment, driver_score, "
            f"manager_rating, manager_comment, manager_score, "
            f"general_review, processed_date, processing_result, extra_data) VALUES ("
            f"{q(TENANT)}, {q(rid)}, "
            f"{q(to_date_only(r.get('Дата відгуку')))}, "
            f"{q(nn(r.get('Статус відгуку')) or 'new')}, "
            f"{lookup('routes', 'rte_id', r.get('RTE_ID'))}, "
            f"{lookup('passengers', 'pax_id', r.get('PAX_ID'))}, "
            f"{lookup('clients_directory', 'cli_id', cli_id)}, "
            f"{q(norm_phone(r.get('Телефон клієнта')))}, {q(nn(r.get('Піб клієнта')))}, "
            f"{q(nn(r.get('Тип запису')))}, "
            f"{q(to_rating(r.get('Оцінка водія')))}, {q(nn(r.get('Коментар про водія')))}, "
            f"{q(to_num(r.get('Бал водія')))}, "
            f"{q(to_rating(r.get('Оцінка менеджера')))}, {q(nn(r.get('Коментар про менеджера')))}, "
            f"{q(to_num(r.get('Бал менеджера')))}, "
            f"{q(nn(r.get('Загальний відгук')))}, "
            f"{q(to_date_any(r.get('Дата опрацювання')))}, "
            f"{q(nn(r.get('Результат')))}, "
            f"{q_jsonb(extra)}"
            f") ON CONFLICT (tenant_id, review_id) WHERE tenant_id = 'esco' DO NOTHING;"
        )
    return sql


def build_client_ratings(rows):
    sql = []
    for r in rows:
        rid = nn(r.get('RATE_ID'))
        if not rid: continue
        cli_id = nn(r.get('CLIENT_ID'))
        if not cli_id: continue  # client_id NOT NULL
        rdate = to_date_only(r.get('Дата оцінки'))
        if not rdate:
            continue  # rating_date NOT NULL
        extra = {
            'route_date': r.get('Дата рейсу'),
        }
        sql.append(
            f"INSERT INTO public.client_ratings ("
            f"tenant_id, rate_id, rating_date, client_id, client_phone, client_name, "
            f"route_id, passenger_id, record_type, "
            f"driver_rating, driver_comment, driver_name, "
            f"manager_rating, manager_comment, manager_name, extra_data) VALUES ("
            f"{q(TENANT)}, {q(rid)}, {q(rdate)}, "
            f"{lookup('clients_directory', 'cli_id', cli_id)}, "
            f"{q(norm_phone(r.get('Телефон клієнта')))}, {q(nn(r.get('Піб клієнта')))}, "
            f"{lookup('routes', 'rte_id', r.get('RTE_ID'))}, "
            f"{lookup('passengers', 'pax_id', r.get('PAX_ID'))}, "
            f"{q(nn(r.get('Тип запису')))}, "
            f"{q(to_rating(r.get('Оцінка водія')))}, {q(nn(r.get('Коментар водія')))}, "
            f"{q(nn(r.get('Водій')))}, "
            f"{q(to_rating(r.get('Оцінка менеджера')))}, {q(nn(r.get('Коментар менеджера')))}, "
            f"{q(nn(r.get('Менеджер')))}, "
            f"{q_jsonb(extra)}"
            f") ON CONFLICT (tenant_id, rate_id) WHERE tenant_id = 'esco' DO NOTHING;"
        )
    return sql


def main():
    msgs = read_sheet(XLSX_KLI, 'Чат')
    racc = read_sheet(XLSX_CONF, 'Маршрути_доступ')
    capp = read_sheet(XLSX_CONF, 'Клієнти_доступ')
    revs = read_sheet(XLSX_KLI, 'Відгуки клієнтів')
    rats = read_sheet(XLSX_KLI, 'Рейтинг клієнтів')

    msg_sql = build_messages(msgs)
    racc_sql = build_route_access(racc)
    capp_sql = build_client_app_access(capp)
    revs_sql = build_reviews(revs)
    rats_sql = build_client_ratings(rats)

    out = []
    out.append("-- =========================================================================")
    out.append("-- ESCO MIGRATION — PHASE 6 (FINAL): MISC")
    out.append("-- =========================================================================")
    out.append(f"-- messages:           {len(msg_sql)}")
    out.append(f"-- route_access:       {len(racc_sql)}")
    out.append(f"-- client_app_access:  {len(capp_sql)}  (без password_hash)")
    out.append(f"-- reviews:            {len(revs_sql)}")
    out.append(f"-- client_ratings:     {len(rats_sql)}")
    out.append(f"-- ────────────────────────")
    out.append(f"-- total:              {len(msg_sql)+len(racc_sql)+len(capp_sql)+len(revs_sql)+len(rats_sql)}")
    out.append("-- =========================================================================")
    out.append("")
    out.append("BEGIN;")
    out.append("")
    out.append("-- ── Schema changes ──────────────────────────────────────────────")
    out.append("ALTER TABLE public.route_access ALTER COLUMN staff_id DROP NOT NULL;")
    out.append("ALTER TABLE public.route_access ALTER COLUMN route_id DROP NOT NULL;")
    out.append("ALTER TABLE public.reviews ADD COLUMN IF NOT EXISTS extra_data jsonb;")
    out.append("ALTER TABLE public.client_ratings ADD COLUMN IF NOT EXISTS extra_data jsonb;")
    out.append("")
    out.append("-- ── Partial UNIQUE indexes (WHERE tenant='esco') ────────────────")
    out.append("CREATE UNIQUE INDEX IF NOT EXISTS messages_esco_msg_uidx ON public.messages (tenant_id, message_id) WHERE tenant_id = 'esco';")
    out.append("CREATE UNIQUE INDEX IF NOT EXISTS route_access_esco_acc_uidx ON public.route_access (tenant_id, access_id) WHERE tenant_id = 'esco';")
    out.append("CREATE UNIQUE INDEX IF NOT EXISTS client_app_access_esco_acc_uidx ON public.client_app_access (tenant_id, access_id) WHERE tenant_id = 'esco';")
    out.append("CREATE UNIQUE INDEX IF NOT EXISTS reviews_esco_rev_uidx ON public.reviews (tenant_id, review_id) WHERE tenant_id = 'esco';")
    out.append("CREATE UNIQUE INDEX IF NOT EXISTS client_ratings_esco_rat_uidx ON public.client_ratings (tenant_id, rate_id) WHERE tenant_id = 'esco';")
    out.append("")
    out.append(f"-- ── messages ({len(msg_sql)}) ─────────────────────────────────────")
    out.extend(msg_sql)
    out.append("")
    out.append(f"-- ── route_access ({len(racc_sql)}) ────────────────────────────────")
    out.extend(racc_sql)
    out.append("")
    out.append(f"-- ── client_app_access ({len(capp_sql)}) ───────────────────────────")
    out.extend(capp_sql)
    out.append("")
    out.append(f"-- ── reviews ({len(revs_sql)}) ──────────────────────────────────────")
    out.extend(revs_sql)
    out.append("")
    out.append(f"-- ── client_ratings ({len(rats_sql)}) ──────────────────────────────")
    out.extend(rats_sql)
    out.append("")
    out.append("COMMIT;")
    out.append("")
    out.append("NOTIFY pgrst, 'reload schema';")

    with open(OUT, 'w', encoding='utf-8') as f:
        f.write('\n'.join(out))

    total = len(msg_sql)+len(racc_sql)+len(capp_sql)+len(revs_sql)+len(rats_sql)
    print(f'✅ {OUT}')
    print(f'   messages:          {len(msg_sql)}')
    print(f'   route_access:      {len(racc_sql)}')
    print(f'   client_app_access: {len(capp_sql)}')
    print(f'   reviews:           {len(revs_sql)}')
    print(f'   client_ratings:    {len(rats_sql)}')
    print(f'   ──────────────')
    print(f'   total:             {total}')


if __name__ == '__main__':
    main()
